import pandas as pd
from datetime import datetime, timedelta

from openpyxl import load_workbook
wb = load_workbook('Reports.xlsx')
print(wb.sheetnames)



def split_date(A):
    df = []
    for i in range(0, len(A)):
        Date = A['Dates:'][i]
        Begin = Date.split('-')[0].strip()
        End = Date.split('-')[1].strip()
        Year_End = int(End.split('/')[-1])
        if int(Begin.split('/')[0]) > int(End.split('/')[0]):
            Year_Begin = Year_End - 1
        else:
            Year_Begin = Year_End
        Begin += '/'
        Begin += str(Year_Begin)
        B = A['Details:'][i]
        for j in ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']:
            if j in A['Details:'][i]:
                B = B.replace(j, '?')
        if '(not submitted)' not in B:
            B = B.split('?')[1:]
        else:
            B = [B]
        df.append(pd.DataFrame({'Begin:': Begin, 'End:': End, 'Details:': B}))
    Frame = pd.concat(df)
    Order = ['Begin:', 'End:', 'Details:']
    Frame = Frame[Order].reset_index(drop=True)        
    return Frame  

def detail_date(A):
    df = []
    C = split_date(A)
    for i in range(0, len(C)):
        Begin = C['Begin:'][i]
        End = C['End:'][i]
        Date = C['Details:'][i].split(':')[0].strip()
        if Date != '(not submitted)':
            Detail = C['Details:'][i].split(':',1)[1].strip()
            if Date.split('/')[0] == Begin.split('/')[0]:
                Date += '/'
                Date += Begin.split('/')[-1]
            else:
                Date += '/'
                Date += End.split('/')[-1]
            if len(Date.split('/')[0]) == 1:
                Month = '0'+ Date.split('/')[0]
            else:
                Month = Date.split('/')[0]
            if len(Date.split('/')[1]) == 1:
                Day = '0' + Date.split('/')[1]
            else:
                Day = Date.split('/')[1]
            Year = '20' + Date.split('/')[2]
            Date = Month + '/' + Day + '/' +Year
        else:
            Detail = '(not submitted)'
        if Detail in ['(not submitted)', '(none)', 'Vacation', 'LOA', 'Sick Day']:
            Hours = int(0)
            Begin_time = Detail
            End_time = Detail
        else:
            Time = Detail.split(',')
            Time = [i.strip().split(' - ') for i in Time]
            Hour = []
            for time in Time:                
                if time[0][-2] == 'a':
                    if time[0].split(':')[0] == '12':
                        time1 = '0' + ':' 
                        time1 += time[0].split(':')[1][0:2]
                    else:
                        time1 = time[0][0:-2]
                else:
                    if time[0].split(':')[0] == '12':
                        time1 = time[0][0:-2]
                    else:
                        time1 = str(int(time[0].split(':')[0])+12) + ':'
                        time1 += time[0].split(':')[1][0:2]
                if time[1][-2] == 'a':
                    if time[1].split(':')[0] == '12':
                        time2 = '24' + ':' 
                        time2 += time[1].split(':')[1][0:2]
                    else:
                        time2 = time[1][0:-2]
                else:
                    if time[1].split(':')[0] == '12':
                        time2 = time[1][0:-2]
                    else:
                        time2 = str(int(time[1].split(':')[0])+12) + ':'
                        time2 += time[1].split(':')[1][0:2]                
                hour = int(time2.split(':')[0]) - int(time1.split(':')[0]) + int(time2.split(':')[1])/60 - int(time1.split(':')[1])/60
                Hour.append(hour)
            Hours = sum(Hour) 
            Begin_time = Time[0][0]
            End_time = Time[-1][1]        
        df.append(pd.DataFrame({'Begin:': Begin, 'End:': End, 'Date:': Date, 'Detail:': Detail, 'Hours:': Hours, 'Begin_time': Begin_time, 'End_time': End_time}, index=[0]))
    Frame = pd.concat(df)
    Order = ['Begin:', 'End:', 'Date:', 'Detail:', 'Begin_time', 'End_time', 'Hours:']  
    Frame = Frame[Order].reset_index(drop=True) 
    Break = []
    for i in range(0, len(Frame)-1):
        if Frame['Detail:'][i] in ['(not submitted)', '(none)', 'Vacation', 'LOA', 'Sick Day']:
            Break.append('>24')
        else:
            if Frame['Detail:'][i+1] in ['(not submitted)', '(none)', 'Vacation', 'LOA', 'Sick Day']:
                Break.append('>24')
            else:
                if Frame['End_time'][i][-2] == 'a':
                    if Frame['End_time'][i].split(':')[0] == '12':
                        time1 = '24' + ':' 
                        time1 += Frame['End_time'][i].split(':')[1][0:2]
                    else:
                        time1 = Frame['End_time'][i][0:-2]
                else:
                    if Frame['End_time'][i].split(':')[0] == '12':
                        time1 = Frame['End_time'][i][0:-2]
                    else:
                        time1 = str(int(Frame['End_time'][i].split(':')[0])+12) + ':'
                        time1 += Frame['End_time'][i].split(':')[1][0:2]                    
                if Frame['Begin_time'][i+1][-2] == 'a':
                    if Frame['Begin_time'][i+1].split(':')[0] == '12':
                        time2 = '0' + ':' 
                        time2 += Frame['Begin_time'][i+1].split(':')[1][0:2]
                    else:
                        time2 = Frame['Begin_time'][i+1][0:-2]
                else:
                    if Frame['Begin_time'][i+1].split(':')[0] == '12':
                        time2 = Frame['Begin_time'][i+1][0:-2]
                    else:
                        time2 = str(int(Frame['Begin_time'][i+1].split(':')[0])+12) + ':'
                        time2 += Frame['Begin_time'][i+1].split(':')[1][0:2]                     
                Break_time = int(time2.split(':')[0]) - int(time1.split(':')[0]) + int(24) + int(time2.split(':')[1])/60 - int(time1.split(':')[1])/60   
                Break.append(Break_time)
    if Frame['Detail:'][len(Frame)-1] in ['(not submitted)', '(none)', 'Vacation', 'LOA', 'Sick Day']:
        Break.append('>24')
    else:
        Break.append('unknown')
    se = pd.Series(Break)
    Frame['Break_time'] = se.values
    return Frame      
    
def date_hours(subsheet, time_range):
    A = pd.read_excel('Reports.xlsx', sheet_name = subsheet)
    df = detail_date(A)
    Begin_Date = time_range.split('-')[0].strip()
    if len(Begin_Date.split('/')[0]) == 1:        
        Begin_Month = '0' + Begin_Date.split('/')[0]
    else:
        Begin_Month = Begin_Date.split('/')[0]
    if len(Begin_Date.split('/')[1]) == 1:        
        Begin_Day = '0' + Begin_Date.split('/')[1]
    else:
        Begin_Day = Begin_Date.split('/')[1] 
    Begin_Year = '20' + Begin_Date.split('/')[2]
    Begin_Date = Begin_Month + '/' + Begin_Day + '/' + Begin_Year
    Begin = datetime.strptime(Begin_Date, "%m/%d/%Y")
    End_Date = time_range.split('-')[1].strip()
    if len(End_Date.split('/')[0]) == 1:        
        End_Month = '0' + End_Date.split('/')[0]
    else:
        End_Month = End_Date.split('/')[0]
    if len(End_Date.split('/')[1]) == 1:        
        End_Day = '0' + End_Date.split('/')[1]
    else:
        End_Day = End_Date.split('/')[1] 
    End_Year = '20' + End_Date.split('/')[2]
    End_Date = End_Month + '/' + End_Day + '/' + End_Year
    End = datetime.strptime(End_Date, "%m/%d/%Y")
    Hour = []
    for i in range(0, len(df)):
        if df['Date:'][i] != '(not submitted)':
            if datetime.strptime(df['Date:'][i], "%m/%d/%Y") >= Begin and datetime.strptime(df['Date:'][i], "%m/%d/%Y") <= End:
                Hour.append(df['Hours:'][i])
    Hours = sum(Hour)
    return Hours


schedules_181_187 = pd.read_excel('Reports.xlsx', sheet_name = '181-187')
schedules_191_196 = pd.read_excel('Reports.xlsx', sheet_name = '191-196')
schedules_201_206 = pd.read_excel('Reports.xlsx', sheet_name = '201-206')

schedules_181_187_1 = schedules_181_187[['Schedule 2015/2016: Dates']]['Schedule 2015/2016: Dates'].tolist()
schedules_181_187_2 = schedules_181_187[['Schedule 2016/2017: Dates']]['Schedule 2016/2017: Dates'].tolist()
schedules_181_187_3 = schedules_181_187[['Schedule 2017/2018: Dates']]['Schedule 2017/2018: Dates'].tolist()

schedules_191_196_1 = schedules_191_196[['Schedule 2016/2017: Dates']]['Schedule 2016/2017: Dates'].tolist()
schedules_191_196_2 = schedules_191_196[['Schedule 2017/2018: Dates']]['Schedule 2017/2018: Dates'].tolist()

schedules_201_206_1 = schedules_201_206[['Schedule 2017/2018: Dates']]['Schedule 2017/2018: Dates'].tolist()

def date_hours_sheet(subsheet, schedules, details):
    A = pd.read_excel('Reports.xlsx', sheet_name = subsheet)
    df = detail_date(A)
    frame = []
    for i in range(0, len(schedules)):
        Begin_Date = schedules[i].split('-')[0].strip()
        if len(Begin_Date.split('/')[0]) == 1:        
            Begin_Month = '0' + Begin_Date.split('/')[0]
        else:
            Begin_Month = Begin_Date.split('/')[0]
        if len(Begin_Date.split('/')[1]) == 1:        
            Begin_Day = '0' + Begin_Date.split('/')[1]
        else:
            Begin_Day = Begin_Date.split('/')[1] 
        Begin_Year = '20' + Begin_Date.split('/')[2]
        Begin_Date = Begin_Month + '/' + Begin_Day + '/' + Begin_Year
        Begin = datetime.strptime(Begin_Date, "%m/%d/%Y")
        End_Date = schedules[i].split('-')[1].strip()
        if len(End_Date.split('/')[0]) == 1:        
            End_Month = '0' + End_Date.split('/')[0]
        else:
            End_Month = End_Date.split('/')[0]
        if len(End_Date.split('/')[1]) == 1:        
            End_Day = '0' + End_Date.split('/')[1]
        else:
            End_Day = End_Date.split('/')[1] 
        End_Year = '20' + End_Date.split('/')[2]
        End_Date = End_Month + '/' + End_Day + '/' + End_Year
        End = datetime.strptime(End_Date, "%m/%d/%Y")
        if 'FMIS' in details[i]:
            if 'day' in details[i] or 'night' in details[i]:
                Begin -= timedelta(days=1)
                End -= timedelta(days=1)
        if 'BEH' in details[i] or 'ED' in details[i]:
            Begin -= timedelta(days=2)
            End -= timedelta(days=2)
        Hour = []
        Break_period = []
        for j in range(0, len(df)):
            if df['Date:'][j] != '(not submitted)':
                if datetime.strptime(df['Date:'][j], "%m/%d/%Y") >= Begin and datetime.strptime(df['Date:'][j], "%m/%d/%Y") <= End:
                    Hour.append(df['Hours:'][j])
                    Break_period.append(df['Break_time'][j])
        Hours = sum(Hour)
        if Hours <= 160 and len([i for i in Break_period if type(i)==int]) <= 12:
            if [i for i in Break_period if type(i)!=str] != []:
                if min([i for i in Break_period if type(i)!=str]) >= 10:
                    Violation = 'No'
                else:
                    Violation = 'Yes'
            else:
                Violation = 'No'                    
        else:
            Violation = 'Yes'
        frame.append(pd.DataFrame({'Schedules': schedules[i], 'Hours': Hours, 'Violation': Violation}, index = [0]))
    Frame = pd.concat(frame)
    Order = ['Schedules','Hours', 'Violation']
    Frame = Frame[Order].reset_index(drop=True)
    return Frame

frame_181_1 = date_hours_sheet('181', schedules_181_187_1, schedules_181_187['181/15'])
frame_181_2 = date_hours_sheet('181', schedules_181_187_2, schedules_181_187['181/16'])
frame_181_3 = date_hours_sheet('181', schedules_181_187_3, schedules_181_187['181/17'])

frame_182_1 = date_hours_sheet('182', schedules_181_187_1, schedules_181_187['182/15'])
frame_182_2 = date_hours_sheet('182', schedules_181_187_2, schedules_181_187['182/16'])
frame_182_3 = date_hours_sheet('182', schedules_181_187_3, schedules_181_187['182/17'])

frame_183_1 = date_hours_sheet('183', schedules_181_187_1, schedules_181_187['183/15'])
frame_183_2 = date_hours_sheet('183', schedules_181_187_2, schedules_181_187['183/16'])
frame_183_3 = date_hours_sheet('183', schedules_181_187_3, schedules_181_187['183/17'])

frame_184_1 = date_hours_sheet('184', schedules_181_187_1, schedules_181_187['184/15'])
frame_184_2 = date_hours_sheet('184', schedules_181_187_2, schedules_181_187['184/16'])
frame_184_3 = date_hours_sheet('184', schedules_181_187_3, schedules_181_187['184/17'])

frame_185_1 = date_hours_sheet('185', schedules_181_187_1, schedules_181_187['185/15'])
frame_185_2 = date_hours_sheet('185', schedules_181_187_2, schedules_181_187['185/16'])
frame_185_3 = date_hours_sheet('185', schedules_181_187_3, schedules_181_187['185/17'])

frame_186_1 = date_hours_sheet('186', schedules_181_187_1, schedules_181_187['186/15'])
frame_186_2 = date_hours_sheet('186', schedules_181_187_2, schedules_181_187['186/16'])
frame_186_3 = date_hours_sheet('186', schedules_181_187_3, schedules_181_187['186/17'])

frame_187_2 = date_hours_sheet('187', schedules_181_187_2, schedules_181_187['187/16'])
frame_187_3 = date_hours_sheet('187', schedules_181_187_3, schedules_181_187['187/17'])
            
frame_191_1 = date_hours_sheet('191', schedules_191_196_1, schedules_191_196['191/16'])
frame_191_2 = date_hours_sheet('191', schedules_191_196_2, schedules_191_196['191/17'])

frame_192_1 = date_hours_sheet('192', schedules_191_196_1, schedules_191_196['192/16'])
frame_192_2 = date_hours_sheet('192', schedules_191_196_2, schedules_191_196['192/17'])

frame_193_1 = date_hours_sheet('193', schedules_191_196_1, schedules_191_196['193/16'])
frame_193_2 = date_hours_sheet('193', schedules_191_196_2, schedules_191_196['193/17'])

frame_194_1 = date_hours_sheet('194', schedules_191_196_1, schedules_191_196['194/16'])
frame_194_2 = date_hours_sheet('194', schedules_191_196_2, schedules_191_196['194/17'])

frame_195_1 = date_hours_sheet('195', schedules_191_196_1, schedules_191_196['195/16'])
frame_195_2 = date_hours_sheet('195', schedules_191_196_2, schedules_191_196['195/17'])

frame_196_1 = date_hours_sheet('196', schedules_191_196_1, schedules_191_196['196/16'])
frame_196_2 = date_hours_sheet('196', schedules_191_196_2, schedules_191_196['196/17'])

frame_201_1 = date_hours_sheet('201', schedules_201_206_1, schedules_201_206['201/17'])

frame_202_1 = date_hours_sheet('202', schedules_201_206_1, schedules_201_206['202/17'])

frame_203_1 = date_hours_sheet('203', schedules_201_206_1, schedules_201_206['203/17'])

frame_204_1 = date_hours_sheet('204', schedules_201_206_1, schedules_201_206['204/17'])

frame_205_1 = date_hours_sheet('205', schedules_201_206_1, schedules_201_206['205/17'])

frame_206_1 = date_hours_sheet('206', schedules_201_206_1, schedules_201_206['206/17'])



detail_date(pd.read_excel('Reports.xlsx', sheet_name = '206')).to_excel('206.xlsx', index = False)




